"""
Sistema de Gestion Clinica - Cuadra Erre (SGCE) v2.0
Flask + PostgreSQL + Email (SMTP)
"""
import re, io, smtplib, ssl, threading, time, os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta, date
from functools import wraps

import schedule
import psycopg2
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from config import EMAIL_REMITENTE, EMAIL_PASSWORD_APP, EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT
except ImportError:
    EMAIL_REMITENTE = ""
    EMAIL_PASSWORD_APP = ""
    EMAIL_SMTP_SERVER = "smtp.gmail.com"
    EMAIL_SMTP_PORT = 587

app = Flask(__name__)
app.secret_key = "SGCE_CuadraErre_2024_xZ9mKqW_v2"

# ============================================================
#  EXPRESIONES REGULARES
# ============================================================
REGEX = {
    "username":    re.compile(r"^@[a-zA-ZáéíóúÁÉÍÓÚñÑ]{2,30}#$"),
    "password":    re.compile(r"^#[a-zA-Z]{3}[0-9]{1,10}[.!@$%&*\-_]+$"),
    "nombre":      re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s'\-]{2,80}$"),
    "apellido":    re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s'\-]{2,80}$"),
    "email":       re.compile(r"^[\w.\-+]+@[\w\-]+\.[a-zA-Z]{2,10}$"),
    "telefono":    re.compile(r"^[\+]?[0-9]{7,15}$"),
    "user_create": re.compile(r"^[a-zA-Z0-9_.\-]{3,30}$"),
    "diagnostico": re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ0-9\s,.\-/()\n]{3,500}$"),
    "alergias":    re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ0-9\s,.\-/()]{0,300}$"),
    "medicamentos":re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ0-9\s,.\-/()mg]{0,300}$"),
    "fecha":       re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    "fecha_hora":  re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}$"),
    "eq_nombre":   re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s'\-]{2,80}$"),
    "raza":        re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s'\-]{0,80}$"),
    "color":       re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s'\-]{0,50}$"),
    "edad":        re.compile(r"^[0-9]{1,2}$"),
    "notas":       re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ0-9\s,.\-/()!?:;\n]{0,500}$"),
    "objetivos":   re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ0-9\s,.\-/()!?:;\n]{0,500}$"),
    "duracion":    re.compile(r"^[0-9]{1,3}$"),
    "parentesco":  re.compile(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]{2,50}$"),
}

MSGS = {
    "username":    "Usuario invalido - formato: @nombre#",
    "password":    "Contrasena invalida - formato: #abc123.",
    "nombre":      "Nombre invalido - solo letras y espacios (2-80 caracteres)",
    "apellido":    "Apellido invalido - solo letras y espacios (2-80 caracteres)",
    "email":       "Correo invalido - ejemplo: nombre@dominio.com",
    "telefono":    "Telefono invalido - 7-15 digitos",
    "user_create": "Usuario invalido - letras, numeros, puntos y guiones (3-30 caracteres)",
    "diagnostico": "Diagnostico invalido - caracteres no permitidos",
    "alergias":    "Alergias invalidas - caracteres no permitidos",
    "medicamentos":"Medicamentos invalidos - caracteres no permitidos",
    "fecha":       "Fecha invalida - formato AAAA-MM-DD",
    "fecha_hora":  "Fecha/hora invalida",
    "eq_nombre":   "Nombre del equino invalido",
    "raza":        "Raza invalida",
    "color":       "Color invalido",
    "edad":        "Edad invalida - numero de 1-99",
    "notas":       "Notas invalidas - caracteres no permitidos",
    "objetivos":   "Objetivos invalidos - caracteres no permitidos",
    "duracion":    "Duracion invalida - numero de minutos (1-999)",
    "parentesco":  "Parentesco invalido - solo letras (2-50 caracteres)",
}

def val(campo, valor, obligatorio=True):
    if not valor or str(valor).strip() == "":
        return (False, "El campo es obligatorio") if obligatorio else (True, "")
    if campo in REGEX and not REGEX[campo].match(str(valor).strip()):
        return False, MSGS.get(campo, "Formato invalido")
    return True, ""

def val_fields(fields):
    errs = []
    for campo, (valor, req) in fields.items():
        ok, m = val(campo, valor, req)
        if not ok:
            errs.append(m)
    return errs

# ============================================================
#  BASE DE DATOS - PostgreSQL
# ============================================================
DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://postgres:password@localhost/cuadraerre")

def get_db():
    return psycopg2.connect(DATABASE_URL)

def qdb(sql, args=(), one=False):
    conn = get_db(); cur = conn.cursor()
    cur.execute(sql, args)
    cols = [d[0].upper() for d in (cur.description or [])]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    return (rows[0] if rows else None) if one else rows

def edb(sql, args=()):
    conn = get_db(); cur = conn.cursor()
    cur.execute(sql, args)
    conn.commit(); conn.close()

def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS roles (
            id INTEGER NOT NULL, nombre VARCHAR(50) NOT NULL, descripcion VARCHAR(200),
            CONSTRAINT pk_roles PRIMARY KEY (id));
        CREATE SEQUENCE IF NOT EXISTS seq_roles;
        CREATE SEQUENCE IF NOT EXISTS seq_usuarios;
        CREATE SEQUENCE IF NOT EXISTS seq_pacientes;
        CREATE SEQUENCE IF NOT EXISTS seq_familiares;
        CREATE SEQUENCE IF NOT EXISTS seq_equinos;
        CREATE SEQUENCE IF NOT EXISTS seq_areas;
        CREATE SEQUENCE IF NOT EXISTS seq_sesiones;
        CREATE SEQUENCE IF NOT EXISTS seq_registro_equino;
        CREATE SEQUENCE IF NOT EXISTS seq_notificaciones;
        CREATE SEQUENCE IF NOT EXISTS seq_auditoria;
        CREATE SEQUENCE IF NOT EXISTS seq_configuracion;
        CREATE SEQUENCE IF NOT EXISTS seq_mantenimiento;
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER NOT NULL, username VARCHAR(80) NOT NULL, password_hash VARCHAR(255) NOT NULL,
            nombre VARCHAR(100) NOT NULL, apellido VARCHAR(100), email VARCHAR(150), telefono VARCHAR(20),
            rol_id INTEGER NOT NULL, activo SMALLINT DEFAULT 1, fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_usuarios PRIMARY KEY (id), CONSTRAINT uq_username UNIQUE (username),
            CONSTRAINT fk_usu_rol FOREIGN KEY (rol_id) REFERENCES roles(id));
        CREATE TABLE IF NOT EXISTS pacientes (
            id INTEGER NOT NULL, nombre VARCHAR(100) NOT NULL, apellido VARCHAR(100) NOT NULL,
            fecha_nacimiento DATE, diagnostico VARCHAR(500), alergias VARCHAR(300), medicamentos VARCHAR(300),
            contacto_emergencia VARCHAR(150), telefono_emergencia VARCHAR(20), email_familiar VARCHAR(150),
            activo SMALLINT DEFAULT 1, fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_pacientes PRIMARY KEY (id));
        CREATE TABLE IF NOT EXISTS familiares (
            id INTEGER NOT NULL, paciente_id INTEGER NOT NULL, usuario_id INTEGER,
            nombre VARCHAR(100) NOT NULL, apellido VARCHAR(100) NOT NULL, parentesco VARCHAR(50),
            telefono VARCHAR(20), email VARCHAR(150),
            CONSTRAINT pk_familiares PRIMARY KEY (id),
            CONSTRAINT fk_fam_paciente FOREIGN KEY (paciente_id) REFERENCES pacientes(id));
        CREATE TABLE IF NOT EXISTS equinos (
            id INTEGER NOT NULL, nombre VARCHAR(80) NOT NULL, raza VARCHAR(80), edad INTEGER,
            color VARCHAR(50), estado VARCHAR(30) DEFAULT 'DISPONIBLE', notas_salud VARCHAR(500),
            ultima_revision DATE, activo SMALLINT DEFAULT 1,
            CONSTRAINT pk_equinos PRIMARY KEY (id));
        CREATE TABLE IF NOT EXISTS areas (
            id INTEGER NOT NULL, nombre VARCHAR(100) NOT NULL, descripcion VARCHAR(200),
            capacidad INTEGER DEFAULT 1, activo SMALLINT DEFAULT 1,
            CONSTRAINT pk_areas PRIMARY KEY (id));
        CREATE TABLE IF NOT EXISTS sesiones (
            id INTEGER NOT NULL, paciente_id INTEGER NOT NULL, terapeuta_id INTEGER NOT NULL,
            equino_id INTEGER, area_id INTEGER, fecha_hora TIMESTAMP NOT NULL,
            duracion_min INTEGER DEFAULT 45, estado VARCHAR(30) DEFAULT 'PROGRAMADA',
            objetivos VARCHAR(500), notas_sesion VARCHAR(1000), recomendaciones_casa VARCHAR(500),
            estado_paciente VARCHAR(50), area_trabajada VARCHAR(80), email_enviado SMALLINT DEFAULT 0,
            recordatorio_enviado SMALLINT DEFAULT 0, fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_sesiones PRIMARY KEY (id),
            CONSTRAINT fk_ses_paciente FOREIGN KEY (paciente_id) REFERENCES pacientes(id),
            CONSTRAINT fk_ses_terapeuta FOREIGN KEY (terapeuta_id) REFERENCES usuarios(id),
            CONSTRAINT fk_ses_equino FOREIGN KEY (equino_id) REFERENCES equinos(id),
            CONSTRAINT fk_ses_area FOREIGN KEY (area_id) REFERENCES areas(id));
        CREATE TABLE IF NOT EXISTS registro_equino (
            id INTEGER NOT NULL, equino_id INTEGER NOT NULL, fecha DATE NOT NULL,
            encargado_id INTEGER, estado_fisico VARCHAR(50), estado_animo VARCHAR(50),
            disponible SMALLINT DEFAULT 1, observaciones VARCHAR(500),
            CONSTRAINT pk_reg_eq PRIMARY KEY (id),
            CONSTRAINT fk_reg_equino FOREIGN KEY (equino_id) REFERENCES equinos(id));
        CREATE TABLE IF NOT EXISTS notificaciones (
            id INTEGER NOT NULL, sesion_id INTEGER, familiar_id INTEGER, tipo VARCHAR(50),
            mensaje VARCHAR(1000), enviada SMALLINT DEFAULT 0, email_destino VARCHAR(150),
            fecha_envio TIMESTAMP, fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_notif PRIMARY KEY (id),
            CONSTRAINT fk_not_sesion FOREIGN KEY (sesion_id) REFERENCES sesiones(id),
            CONSTRAINT fk_not_fam FOREIGN KEY (familiar_id) REFERENCES familiares(id));
        CREATE TABLE IF NOT EXISTS auditoria (
            id INTEGER NOT NULL, usuario_id INTEGER, accion VARCHAR(100),
            tabla_afectada VARCHAR(50), registro_id INTEGER, detalle VARCHAR(500),
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_auditoria PRIMARY KEY (id));
        CREATE TABLE IF NOT EXISTS configuracion (
            id INTEGER NOT NULL, nombre_centro VARCHAR(150) DEFAULT 'Cuadra Erre',
            telefono VARCHAR(20), direccion VARCHAR(300), email_contacto VARCHAR(150),
            color_primario VARCHAR(10) DEFAULT '#1e3a5f', recordatorio_24h SMALLINT DEFAULT 1,
            horas_anticipacion INTEGER DEFAULT 24,
            CONSTRAINT pk_config PRIMARY KEY (id));
        CREATE TABLE IF NOT EXISTS mantenimiento_equino (
            id INTEGER NOT NULL, equino_id INTEGER NOT NULL, tipo VARCHAR(50) NOT NULL,
            fecha_programada DATE NOT NULL, fecha_realizada DATE, estado VARCHAR(30) DEFAULT 'PENDIENTE',
            notas VARCHAR(500), registrado_por INTEGER, fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT pk_mant_eq PRIMARY KEY (id),
            CONSTRAINT fk_mant_equino FOREIGN KEY (equino_id) REFERENCES equinos(id),
            CONSTRAINT fk_mant_usuario FOREIGN KEY (registrado_por) REFERENCES usuarios(id));
    """)
    # Insertar datos iniciales solo si las tablas estan vacias
    cur.execute("SELECT COUNT(*) FROM roles")
    if cur.fetchone()[0] == 0:
        cur.execute("""
            INSERT INTO roles (id, nombre, descripcion) VALUES
            (nextval('seq_roles'),'ADMINISTRADOR','Acceso total al sistema'),
            (nextval('seq_roles'),'TERAPEUTA','Gestion de sesiones y expedientes'),
            (nextval('seq_roles'),'COORDINADOR','Agenda y reportes'),
            (nextval('seq_roles'),'ENCARGADO_EQUINOS','Registro de caballos');
            INSERT INTO areas (id, nombre, descripcion, capacidad) VALUES
            (nextval('seq_areas'),'Picadero 1','Pista principal',1),
            (nextval('seq_areas'),'Picadero 2','Pista secundaria',1),
            (nextval('seq_areas'),'Sala de Evaluacion','Evaluacion inicial',2);
            INSERT INTO equinos (id, nombre, raza, edad, color, estado) VALUES
            (nextval('seq_equinos'),'Tornado','Cuarto de Milla',8,'Bayo','DISPONIBLE'),
            (nextval('seq_equinos'),'Luna','Andaluz',6,'Blanco','DISPONIBLE'),
            (nextval('seq_equinos'),'Rayo','Mestizo',10,'Negro','DISPONIBLE'),
            (nextval('seq_equinos'),'Estrella','Pura Sangre',7,'Alazan','DISPONIBLE'),
            (nextval('seq_equinos'),'Oso','Azteca',5,'Castano','DISPONIBLE');
            INSERT INTO pacientes (id, nombre, apellido, fecha_nacimiento, diagnostico) VALUES
            (nextval('seq_pacientes'),'Sofia','Garcia','2015-03-12','Trastorno del Espectro Autista'),
            (nextval('seq_pacientes'),'Miguel','Torres','2012-07-08','TDAH'),
            (nextval('seq_pacientes'),'Lucas','Perez','2014-05-30','Sindrome de Down');
            INSERT INTO configuracion (id, nombre_centro, telefono, direccion, email_contacto, color_primario, recordatorio_24h, horas_anticipacion)
            VALUES (nextval('seq_configuracion'),'Cuadra Erre','624-000-0000','Los Cabos, Baja California Sur','cuadraerreoficial@gmail.com','#1e3a5f',1,24);
        """)
        admin_hash = generate_password_hash("#Adm123.")
        cur.execute(
            "INSERT INTO usuarios (id, username, password_hash, nombre, apellido, email, rol_id) "
            "VALUES (nextval('seq_usuarios'),'admin',%s,'Administrador','Sistema','admin@cuadraerre.mx',1)",
            (admin_hash,)
        )
    conn.commit(); conn.close()

try:
    init_db()
except Exception as e:
    print(f"[init_db] Error: {e}")

def registrar_auditoria(accion, tabla, registro_id, detalle=""):
    try:
        uid = session.get("uid")
        edb(
            "INSERT INTO AUDITORIA(ID,USUARIO_ID,ACCION,TABLA_AFECTADA,REGISTRO_ID,DETALLE,FECHA) "
            "VALUES(nextval('seq_auditoria'),%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)",
            (uid, accion, tabla, registro_id, detalle[:500])
        )
    except Exception:
        pass

def generar_excel(titulo, headers, rows, nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    AZUL_OSC = "1E3A5F"
    GRIS_CLR = "F4F6F8"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell_t = ws.cell(row=1, column=1, value="CUADRA ERRE - " + titulo)
    cell_t.font = Font(bold=True, size=14, color="FFFFFF")
    cell_t.fill = PatternFill("solid", fgColor=AZUL_OSC)
    cell_t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    cell_s = ws.cell(row=2, column=1, value=f"Generado el {date.today().strftime('%d/%m/%Y')}")
    cell_s.font = Font(italic=True, size=9, color="5A6678")

    header_row = 4
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_OSC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="D4D9E0"))

    for r_idx, row in enumerate(rows, start=header_row+1):
        fill = PatternFill("solid", fgColor=GRIS_CLR) if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.border = Border(bottom=Side(style="thin", color="E2E6EA"))
            cell.alignment = Alignment(vertical="center")

    for idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in rows:
            val = row[idx-1] if idx-1 < len(row) else ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nombre_archivo)

# ============================================================
#  ENVIO DE EMAIL
# ============================================================
def enviar_email(destinatario, asunto, cuerpo_html):
    if not EMAIL_REMITENTE or not EMAIL_PASSWORD_APP:
        return False, "Email no configurado"
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = asunto
        msg["From"] = f"Cuadra Erre <{EMAIL_REMITENTE}>"
        msg["To"] = destinatario
        msg.attach(MIMEText(cuerpo_html, "html", "utf-8"))

        context = ssl.create_default_context()
        with smtplib.SMTP(EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(EMAIL_REMITENTE, EMAIL_PASSWORD_APP)
            server.sendmail(EMAIL_REMITENTE, destinatario, msg.as_string())
        return True, "Enviado"
    except Exception as e:
        return False, str(e)

def plantilla_email(titulo, cuerpo_texto, color_acento="#1e3a5f"):
    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;max-width:560px;margin:0 auto;
                border:1px solid #d4d9e0;border-radius:6px;overflow:hidden">
      <div style="background:{color_acento};padding:20px 28px">
        <span style="color:#fff;font-size:18px;font-weight:600;letter-spacing:.3px">CUADRA ERRE</span>
        <div style="color:#c5d2e3;font-size:11px;margin-top:2px">Sistema de Gestion Clinica</div>
      </div>
      <div style="padding:28px;background:#ffffff">
        <h2 style="color:#1a2942;font-size:16px;margin:0 0 14px;font-weight:600">{titulo}</h2>
        <div style="color:#3d4a5c;font-size:14px;line-height:1.6">{cuerpo_texto}</div>
      </div>
      <div style="background:#f4f6f8;padding:14px 28px;border-top:1px solid #e2e6ea">
        <span style="color:#8a94a3;font-size:11px">
          Este es un mensaje automatico del Sistema de Gestion Clinica Cuadra Erre, Los Cabos BCS.
          Por favor no responda a este correo.
        </span>
      </div>
    </div>
    """

# ============================================================
#  CONFIGURACION DEL SISTEMA
# ============================================================
def get_config():
    cfg = qdb("SELECT * FROM CONFIGURACION LIMIT 1", one=True)
    if not cfg:
        return {
            "NOMBRE_CENTRO": "Cuadra Erre", "TELEFONO": "", "DIRECCION": "",
            "EMAIL_CONTACTO": EMAIL_REMITENTE, "COLOR_PRIMARIO": "#1e3a5f",
            "RECORDATORIO_24H": 1, "HORAS_ANTICIPACION": 24
        }
    return cfg

# ============================================================
#  RECORDATORIOS AUTOMATICOS (hilo en segundo plano)
# ============================================================
def enviar_recordatorios():
    try:
        cfg = get_config()
        if not cfg or int(cfg.get("RECORDATORIO_24H", 1)) != 1:
            return
        horas = int(cfg.get("HORAS_ANTICIPACION", 24))

        ahora = datetime.now()
        limite = ahora + timedelta(hours=horas)

        sesiones = qdb(
            "SELECT S.ID, S.FECHA_HORA, P.ID PID, P.NOMBRE||' '||P.APELLIDO PACIENTE, "
            "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, E.NOMBRE EQUINO "
            "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
            "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
            "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
            "WHERE S.ESTADO='PROGRAMADA' AND S.RECORDATORIO_ENVIADO=0 "
            "AND S.FECHA_HORA <= %s AND S.FECHA_HORA >= %s",
            (limite, ahora)
        )
        for s in (sesiones or []):
            familiares = qdb(
                "SELECT ID, NOMBRE, EMAIL FROM FAMILIARES WHERE PACIENTE_ID=%s AND EMAIL IS NOT NULL",
                (s["PID"],)
            )
            fecha_fmt = str(s["FECHA_HORA"])[:16]
            for fam in (familiares or []):
                cuerpo = f"""
                <p>Estimado(a) {fam['NOMBRE']},</p>
                <p>Le recordamos que <strong>{s['PACIENTE']}</strong> tiene una sesion
                de equinoterapia programada proximamente:</p>
                <table style="width:100%;border-collapse:collapse;margin:14px 0">
                  <tr><td style="padding:6px 0;color:#5a6678">Fecha y hora</td>
                      <td style="padding:6px 0;font-weight:600;color:#1a2942">{fecha_fmt}</td></tr>
                  <tr><td style="padding:6px 0;color:#5a6678">Terapeuta</td>
                      <td style="padding:6px 0;font-weight:600;color:#1a2942">{s['TERAPEUTA']}</td></tr>
                  <tr><td style="padding:6px 0;color:#5a6678">Equino asignado</td>
                      <td style="padding:6px 0;font-weight:600;color:#1a2942">{s['EQUINO'] or 'Por asignar'}</td></tr>
                </table>
                <p>Le pedimos llegar 10 minutos antes de la hora programada.</p>
                """
                ok, _ = enviar_email(fam["EMAIL"], "Recordatorio de sesion - Cuadra Erre",
                                     plantilla_email("Recordatorio de sesion proxima", cuerpo))
                notif_id = qdb("SELECT nextval('seq_notificaciones') AS id", one=True)["ID"]
                edb(
                    "INSERT INTO NOTIFICACIONES(ID,SESION_ID,FAMILIAR_ID,TIPO,MENSAJE,ENVIADA,EMAIL_DESTINO,FECHA_ENVIO,FECHA_CREACION) "
                    "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)",
                    (notif_id, s["ID"], fam["ID"], "RECORDATORIO",
                     f"Recordatorio enviado para sesion {fecha_fmt}", 1 if ok else 0, fam["EMAIL"],
                     datetime.now() if ok else None)
                )
            edb("UPDATE SESIONES SET RECORDATORIO_ENVIADO=1 WHERE ID=%s", (s["ID"],))
    except Exception as e:
        print(f"[Recordatorios] Error: {e}")

def iniciar_scheduler():
    schedule.every(15).minutes.do(enviar_recordatorios)
    while True:
        schedule.run_pending()
        time.sleep(30)

scheduler_thread = threading.Thread(target=iniciar_scheduler, daemon=True)
scheduler_thread.start()

# ============================================================
#  DECORADORES
# ============================================================
def login_req(f):
    @wraps(f)
    def d(*a, **kw):
        if "uid" not in session:
            flash("Debes iniciar sesion", "warning")
            return redirect(url_for("login"))
        return f(*a, **kw)
    return d

def rol_req(*roles):
    def dec(f):
        @wraps(f)
        def d(*a, **kw):
            if session.get("rol") not in roles:
                flash("Sin permiso para esta accion", "danger")
                return redirect(url_for("dashboard"))
            return f(*a, **kw)
        return d
    return dec

# ============================================================
#  LOGIN
# ============================================================
@app.route("/", methods=["GET", "POST"])
def login():
    if "uid" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        ok_u, msg_u = val("username", username)
        if not ok_u:
            flash(msg_u, "danger")
            return render_template("login.html")

        ok_p, msg_p = val("password", password)
        if not ok_p:
            flash(msg_p, "danger")
            return render_template("login.html")

        nombre_ingresado = username[1:-1].lower()
        letras_pass      = password[1:4].lower()

        usuarios = qdb(
            "SELECT U.ID, U.USERNAME, U.NOMBRE, U.APELLIDO, U.PASSWORD_HASH, "
            "R.NOMBRE ROL FROM USUARIOS U JOIN ROLES R ON U.ROL_ID=R.ID "
            "WHERE U.ACTIVO=1"
        )
        for u in (usuarios or []):
            nombre_bd = u["NOMBRE"].lower()
            if nombre_bd.startswith(nombre_ingresado):
                if nombre_bd[:3] == letras_pass:
                    if check_password_hash(u["PASSWORD_HASH"], password):
                        session["uid"]    = u["ID"]
                        session["nombre"] = f"{u['NOMBRE']} {u['APELLIDO'] or ''}".strip()
                        session["user"]   = u["USERNAME"]
                        session["rol"]    = u["ROL"]
                        registrar_auditoria("LOGIN", "USUARIOS", u["ID"], f"Inicio de sesion: {u['USERNAME']}")
                        flash(f"Bienvenido, {session['nombre']}", "success")
                        return redirect(url_for("dashboard"))
                    else:
                        flash("Contrasena incorrecta", "danger")
                        return render_template("login.html")
                else:
                    flash("Contrasena incorrecta", "danger")
                    return render_template("login.html")

        flash("Usuario no encontrado en el sistema", "danger")
    return render_template("login.html")

@app.route("/logout", methods=["GET", "POST"])
def logout():
    if "uid" in session:
        registrar_auditoria("LOGOUT", "USUARIOS", session["uid"], "Cierre de sesion")
    session.clear()
    return redirect(url_for("login"))

# ============================================================
#  DASHBOARD
# ============================================================
@app.route("/dashboard")
@login_req
def dashboard():
    hoy = date.today().isoformat()
    rol = session["rol"]
    stats = {}

    if rol in ("ADMINISTRADOR", "COORDINADOR"):
        stats["sesiones_hoy"]  = (qdb("SELECT COUNT(*) C FROM SESIONES WHERE CAST(FECHA_HORA AS DATE)=%s", (hoy,), one=True) or {}).get("C", 0)
        stats["pacientes"]     = (qdb("SELECT COUNT(*) C FROM PACIENTES WHERE ACTIVO=1", one=True) or {}).get("C", 0)
        stats["equinos_disp"]  = (qdb("SELECT COUNT(*) C FROM EQUINOS WHERE ESTADO='DISPONIBLE' AND ACTIVO=1", one=True) or {}).get("C", 0)
        stats["equinos_total"] = (qdb("SELECT COUNT(*) C FROM EQUINOS WHERE ACTIVO=1", one=True) or {}).get("C", 0)
        stats["sesiones_semana"] = (qdb(
            "SELECT COUNT(*) C FROM SESIONES WHERE CAST(FECHA_HORA AS DATE) >= %s",
            ((date.today()-timedelta(days=date.today().weekday())).isoformat(),), one=True) or {}).get("C", 0)
        stats["notif_pendientes"] = (qdb("SELECT COUNT(*) C FROM NOTIFICACIONES WHERE ENVIADA=0", one=True) or {}).get("C", 0)
        stats["pacientes_sin_seguimiento"] = (qdb(
            "SELECT COUNT(*) C FROM PACIENTES P WHERE P.ACTIVO=1 AND P.ID NOT IN "
            "(SELECT PACIENTE_ID FROM SESIONES WHERE FECHA_HORA >= %s AND ESTADO != 'CANCELADA')",
            (datetime.now()-timedelta(days=14),), one=True) or {}).get("C", 0)
        stats["equinos_atencion"] = (qdb(
            "SELECT COUNT(*) C FROM EQUINOS WHERE ESTADO IN ('PRECAUCION','DESCANSO','VETERINARIO') AND ACTIVO=1",
            one=True) or {}).get("C", 0)

    elif rol == "TERAPEUTA":
        stats["mis_hoy"]       = (qdb("SELECT COUNT(*) C FROM SESIONES WHERE TERAPEUTA_ID=%s AND CAST(FECHA_HORA AS DATE)=%s", (session["uid"], hoy), one=True) or {}).get("C", 0)
        stats["mis_pacientes"] = (qdb("SELECT COUNT(DISTINCT PACIENTE_ID) C FROM SESIONES WHERE TERAPEUTA_ID=%s", (session["uid"],), one=True) or {}).get("C", 0)
        stats["mis_semana"] = (qdb(
            "SELECT COUNT(*) C FROM SESIONES WHERE TERAPEUTA_ID=%s AND CAST(FECHA_HORA AS DATE) >= %s",
            (session["uid"], (date.today()-timedelta(days=date.today().weekday())).isoformat()), one=True) or {}).get("C", 0)
        stats["notas_pendientes"] = (qdb(
            "SELECT COUNT(*) C FROM SESIONES WHERE TERAPEUTA_ID=%s AND ESTADO='PROGRAMADA' AND FECHA_HORA < CURRENT_TIMESTAMP",
            (session["uid"],), one=True) or {}).get("C", 0)
        proxima = qdb(
            "SELECT S.ID, S.FECHA_HORA, P.NOMBRE||' '||P.APELLIDO PACIENTE, E.NOMBRE EQUINO "
            "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
            "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
            "WHERE S.TERAPEUTA_ID=%s AND S.ESTADO='PROGRAMADA' AND S.FECHA_HORA >= CURRENT_TIMESTAMP "
            "ORDER BY S.FECHA_HORA LIMIT 1", (session["uid"],), one=True
        )
        stats["proxima_sesion"] = proxima

    elif rol == "ENCARGADO_EQUINOS":
        stats["disponible"]  = (qdb("SELECT COUNT(*) C FROM EQUINOS WHERE ESTADO='DISPONIBLE' AND ACTIVO=1", one=True) or {}).get("C", 0)
        stats["precaucion"]  = (qdb("SELECT COUNT(*) C FROM EQUINOS WHERE ESTADO='PRECAUCION' AND ACTIVO=1", one=True) or {}).get("C", 0)
        stats["descanso"]    = (qdb("SELECT COUNT(*) C FROM EQUINOS WHERE ESTADO IN ('DESCANSO','VETERINARIO') AND ACTIVO=1", one=True) or {}).get("C", 0)
        stats["sin_registro_hoy"] = (qdb(
            "SELECT COUNT(*) C FROM EQUINOS E WHERE E.ACTIVO=1 AND E.ID NOT IN "
            "(SELECT EQUINO_ID FROM REGISTRO_EQUINO WHERE CAST(FECHA AS DATE)=%s)",
            (hoy,), one=True) or {}).get("C", 0)
        equinos_pendientes = qdb(
            "SELECT NOMBRE FROM EQUINOS E WHERE E.ACTIVO=1 AND E.ID NOT IN "
            "(SELECT EQUINO_ID FROM REGISTRO_EQUINO WHERE CAST(FECHA AS DATE)=%s) ORDER BY NOMBRE",
            (hoy,)
        )
        stats["lista_sin_registro"] = [e["NOMBRE"] for e in (equinos_pendientes or [])]

    sesiones = qdb(
        "SELECT S.ID, S.FECHA_HORA, S.ESTADO, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, "
        "E.NOMBRE EQUINO "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "WHERE CAST(S.FECHA_HORA AS DATE)>=%s AND S.ESTADO!='CANCELADA' "
        "ORDER BY S.FECHA_HORA LIMIT 8", (hoy,)
    )
    equinos_sem = qdb("SELECT ID, NOMBRE, ESTADO FROM EQUINOS WHERE ACTIVO=1 ORDER BY NOMBRE LIMIT 12")

    grafica_labels = []
    grafica_data = []
    for i in range(-3, 4):
        d = date.today() + timedelta(days=i)
        c = (qdb("SELECT COUNT(*) C FROM SESIONES WHERE CAST(FECHA_HORA AS DATE)=%s AND ESTADO!='CANCELADA'", (d.isoformat(),), one=True) or {}).get("C", 0)
        etiqueta = d.strftime("%d/%m")
        if i == 0:
            etiqueta += " (hoy)"
        grafica_labels.append(etiqueta)
        grafica_data.append(c)

    return render_template("dashboard.html", stats=stats, sesiones=sesiones,
                           equinos=equinos_sem, grafica_labels=grafica_labels,
                           grafica_data=grafica_data)

# ============================================================
#  AGENDA
# ============================================================
@app.route("/agenda")
@login_req
def agenda():
    filtro_estado = request.args.get("estado", "")
    filtro_busqueda = request.args.get("q", "")

    base_sql = (
        "SELECT S.ID, S.FECHA_HORA, S.ESTADO, S.DURACION_MIN, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, "
        "E.NOMBRE EQUINO, A.NOMBRE AREA "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "LEFT JOIN AREAS A ON S.AREA_ID=A.ID WHERE 1=1 "
    )
    args = []
    if session["rol"] == "TERAPEUTA":
        base_sql += "AND S.TERAPEUTA_ID=%s "
        args.append(session["uid"])
    if filtro_estado:
        base_sql += "AND S.ESTADO=%s "
        args.append(filtro_estado)
    if filtro_busqueda:
        base_sql += "AND (UPPER(P.NOMBRE||' '||P.APELLIDO) LIKE %s OR UPPER(U.NOMBRE||' '||U.APELLIDO) LIKE %s) "
        like = f"%{filtro_busqueda.upper()}%"
        args.extend([like, like])
    base_sql += "ORDER BY S.FECHA_HORA DESC LIMIT 150"

    sesiones = qdb(base_sql, tuple(args))
    pacientes  = qdb("SELECT ID, NOMBRE||' '||APELLIDO NOM FROM PACIENTES WHERE ACTIVO=1 ORDER BY APELLIDO")
    terapeutas = qdb("SELECT U.ID, U.NOMBRE||' '||U.APELLIDO NOM FROM USUARIOS U JOIN ROLES R ON U.ROL_ID=R.ID WHERE R.NOMBRE='TERAPEUTA' AND U.ACTIVO=1 ORDER BY U.NOMBRE")
    equinos    = qdb("SELECT ID, NOMBRE FROM EQUINOS WHERE ACTIVO=1 AND ESTADO='DISPONIBLE' ORDER BY NOMBRE")
    areas      = qdb("SELECT ID, NOMBRE FROM AREAS WHERE ACTIVO=1 ORDER BY NOMBRE")
    return render_template("agenda.html", sesiones=sesiones, pacientes=pacientes,
                           terapeutas=terapeutas, equinos=equinos, areas=areas,
                           filtro_estado=filtro_estado, filtro_busqueda=filtro_busqueda)

@app.route("/api/agenda_semana")
@login_req
def api_agenda_semana():
    inicio_str = request.args.get("inicio", "")
    try:
        inicio = datetime.strptime(inicio_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        inicio = date.today() - timedelta(days=date.today().weekday())
    fin = inicio + timedelta(days=6)

    base_sql = (
        "SELECT S.ID, S.FECHA_HORA, S.ESTADO, S.DURACION_MIN, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, "
        "E.NOMBRE EQUINO, A.NOMBRE AREA "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "LEFT JOIN AREAS A ON S.AREA_ID=A.ID "
        "WHERE CAST(S.FECHA_HORA AS DATE) BETWEEN %s AND %s "
    )
    args = [inicio.isoformat(), fin.isoformat()]
    if session["rol"] == "TERAPEUTA":
        base_sql += "AND S.TERAPEUTA_ID=%s "
        args.append(session["uid"])
    base_sql += "ORDER BY S.FECHA_HORA"

    sesiones = qdb(base_sql, tuple(args))
    for s in sesiones:
        s["FECHA_HORA"] = str(s["FECHA_HORA"])

    return jsonify({
        "inicio": inicio.isoformat(),
        "fin": fin.isoformat(),
        "sesiones": sesiones
    })

@app.route("/agenda/nueva", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def nueva_sesion():
    p = request.form
    errores = val_fields({
        "fecha_hora": (p.get("fecha_hora",""), True),
        "objetivos":  (p.get("objetivos",""),  False),
        "duracion":   (p.get("duracion","45"), True),
    })
    if not p.get("paciente_id"): errores.append("Paciente: obligatorio")
    if not p.get("terapeuta_id") and session["rol"] != "TERAPEUTA":
        errores.append("Terapeuta: obligatorio")

    equino_id = p.get("equino_id") or None
    if equino_id:
        eq = qdb("SELECT ESTADO FROM EQUINOS WHERE ID=%s", (equino_id,), one=True)
        if eq and eq["ESTADO"] not in ("DISPONIBLE", "PRECAUCION"):
            errores.append(f"El equino seleccionado esta en estado '{eq['ESTADO']}' y no puede asignarse")

    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("agenda"))

    tid = p["terapeuta_id"] if session["rol"] != "TERAPEUTA" else session["uid"]
    fecha_sql = p["fecha_hora"].replace("T", " ") + ":00"

    nuevo_id = qdb("SELECT nextval('seq_sesiones') AS id", one=True)["ID"]
    edb(
        "INSERT INTO SESIONES(ID,PACIENTE_ID,TERAPEUTA_ID,EQUINO_ID,AREA_ID,"
        "FECHA_HORA,DURACION_MIN,OBJETIVOS,ESTADO) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,'PROGRAMADA')",
        (nuevo_id, p["paciente_id"], tid, equino_id,
         p.get("area_id") or None, fecha_sql, p.get("duracion","45"), p.get("objetivos",""))
    )
    registrar_auditoria("CREAR", "SESIONES", nuevo_id, f"Nueva sesion programada para {fecha_sql}")

    paciente = qdb("SELECT NOMBRE, APELLIDO FROM PACIENTES WHERE ID=%s", (p["paciente_id"],), one=True)
    familiares = qdb("SELECT NOMBRE, EMAIL FROM FAMILIARES WHERE PACIENTE_ID=%s AND EMAIL IS NOT NULL", (p["paciente_id"],))
    terapeuta_nom = qdb("SELECT NOMBRE, APELLIDO FROM USUARIOS WHERE ID=%s", (tid,), one=True)
    equino_nom = qdb("SELECT NOMBRE FROM EQUINOS WHERE ID=%s", (equino_id,), one=True) if equino_id else None

    fecha_fmt = datetime.strptime(fecha_sql, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y a las %H:%M hrs")
    for fam in (familiares or []):
        cuerpo = f"""
        <p>Estimado(a) {fam['NOMBRE']},</p>
        <p>Le confirmamos que se ha programado una nueva sesion de equinoterapia para
        <strong>{paciente['NOMBRE']} {paciente['APELLIDO']}</strong>:</p>
        <table style="width:100%;border-collapse:collapse;margin:14px 0">
          <tr><td style="padding:6px 0;color:#5a6678">Fecha y hora</td>
              <td style="padding:6px 0;font-weight:600;color:#1a2942">{fecha_fmt}</td></tr>
          <tr><td style="padding:6px 0;color:#5a6678">Terapeuta</td>
              <td style="padding:6px 0;font-weight:600;color:#1a2942">{terapeuta_nom['NOMBRE']} {terapeuta_nom['APELLIDO']}</td></tr>
          <tr><td style="padding:6px 0;color:#5a6678">Equino asignado</td>
              <td style="padding:6px 0;font-weight:600;color:#1a2942">{equino_nom['NOMBRE'] if equino_nom else 'Por asignar'}</td></tr>
        </table>
        <p>Le pedimos llegar 10 minutos antes de la hora programada.</p>
        """
        ok, msg = enviar_email(fam["EMAIL"], "Confirmacion de sesion - Cuadra Erre",
                               plantilla_email("Confirmacion de sesion programada", cuerpo))
        notif_id = qdb("SELECT nextval('seq_notificaciones') AS id", one=True)["ID"]
        fam_row = qdb("SELECT ID FROM FAMILIARES WHERE EMAIL=%s AND PACIENTE_ID=%s", (fam["EMAIL"], p["paciente_id"]), one=True)
        edb(
            "INSERT INTO NOTIFICACIONES(ID,SESION_ID,FAMILIAR_ID,TIPO,MENSAJE,ENVIADA,EMAIL_DESTINO,FECHA_ENVIO,FECHA_CREACION) "
            "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)",
            (notif_id, nuevo_id, fam_row["ID"] if fam_row else None, "CONFIRMACION_CITA",
             f"Confirmacion de cita enviada para {fecha_fmt}", 1 if ok else 0, fam["EMAIL"],
             datetime.now() if ok else None)
        )

    flash("Sesion programada correctamente. Confirmacion enviada a familiares.", "success")
    return redirect(url_for("agenda"))

@app.route("/sesion/<int:sid>", methods=["GET", "POST"])
@login_req
def sesion_detalle(sid):
    def get_sesion():
        return qdb(
            "SELECT S.*, P.NOMBRE||' '||P.APELLIDO PACIENTE, P.ID PID, "
            "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, "
            "E.NOMBRE EQUINO, A.NOMBRE AREA "
            "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
            "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
            "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
            "LEFT JOIN AREAS A ON S.AREA_ID=A.ID "
            "WHERE S.ID=%s", (sid,), one=True
        )
    s = get_sesion()
    if not s: abort(404)

    if request.method == "POST":
        p     = request.form
        notas = p.get("notas_sesion","")
        recom = p.get("recomendaciones_casa","")
        errores = val_fields({
            "notas":     (notas, False),
            "objetivos": (recom, False),
        })
        if errores:
            flash(" | ".join(errores), "danger")
            return redirect(url_for("sesion_detalle", sid=sid))

        edb("UPDATE SESIONES SET NOTAS_SESION=%s,RECOMENDACIONES_CASA=%s,"
            "ESTADO_PACIENTE=%s,AREA_TRABAJADA=%s,ESTADO=%s WHERE ID=%s",
            (notas, recom, p.get("estado_paciente",""), p.get("area_trabajada",""),
             p.get("estado","REALIZADA"), sid))
        registrar_auditoria("ACTUALIZAR", "SESIONES", sid, f"Nota clinica guardada, estado: {p.get('estado')}")

        if p.get("estado") == "REALIZADA":
            fams = qdb("SELECT F.ID, F.NOMBRE, F.EMAIL FROM FAMILIARES F WHERE F.PACIENTE_ID=%s AND F.EMAIL IS NOT NULL", (s["PID"],))
            fecha_fmt = str(s["FECHA_HORA"])[:16]
            for fam in (fams or []):
                cuerpo = f"""
                <p>Estimado(a) {fam['NOMBRE']},</p>
                <p>Le compartimos el resumen de la sesion de <strong>{s['PACIENTE']}</strong>
                realizada el {fecha_fmt}:</p>
                <div style="background:#f4f6f8;border-left:3px solid #1e3a5f;padding:12px 16px;margin:14px 0">
                  <strong style="color:#1a2942">Notas de la sesion</strong>
                  <p style="margin:6px 0 0;color:#3d4a5c">{notas or 'Sin notas registradas'}</p>
                </div>
                <div style="background:#f0f5fa;border-left:3px solid #2c5f8a;padding:12px 16px;margin:14px 0">
                  <strong style="color:#1a2942">Recomendaciones para casa</strong>
                  <p style="margin:6px 0 0;color:#3d4a5c">{recom or 'Sin recomendaciones especiales'}</p>
                </div>
                """
                ok, _ = enviar_email(fam["EMAIL"], f"Resumen de sesion - {s['PACIENTE']}",
                                     plantilla_email("Resumen de sesion completada", cuerpo))
                notif_id = qdb("SELECT nextval('seq_notificaciones') AS id", one=True)["ID"]
                edb(
                    "INSERT INTO NOTIFICACIONES(ID,SESION_ID,FAMILIAR_ID,TIPO,MENSAJE,ENVIADA,EMAIL_DESTINO,FECHA_ENVIO,FECHA_CREACION) "
                    "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)",
                    (notif_id, sid, fam["ID"], "NOTA_SESION",
                     f"Resumen enviado: {notas[:100]}", 1 if ok else 0, fam["EMAIL"],
                     datetime.now() if ok else None)
                )
            flash("Sesion guardada y resumen enviado por correo a familiares", "success")
        else:
            flash("Sesion actualizada correctamente", "success")
        return redirect(url_for("sesion_detalle", sid=sid))

    return render_template("sesion_detalle.html", sesion=get_sesion())

@app.route("/sesion/<int:sid>/cancelar", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR")
def cancelar_sesion(sid):
    edb("UPDATE SESIONES SET ESTADO='CANCELADA' WHERE ID=%s", (sid,))
    registrar_auditoria("CANCELAR", "SESIONES", sid, "Sesion cancelada")
    flash("Sesion cancelada", "warning")
    return redirect(url_for("agenda"))

# ============================================================
#  PACIENTES
# ============================================================
@app.route("/pacientes")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def pacientes():
    busqueda = request.args.get("q", "")
    if busqueda:
        lista = qdb(
            "SELECT * FROM PACIENTES WHERE ACTIVO=1 AND "
            "UPPER(NOMBRE||' '||APELLIDO) LIKE %s ORDER BY APELLIDO, NOMBRE",
            (f"%{busqueda.upper()}%",)
        )
    else:
        lista = qdb("SELECT * FROM PACIENTES WHERE ACTIVO=1 ORDER BY APELLIDO, NOMBRE")
    return render_template("pacientes.html", pacientes=lista, busqueda=busqueda)

@app.route("/pacientes/nuevo", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR")
def nuevo_paciente():
    p = request.form
    errores = val_fields({
        "nombre":      (p.get("nombre",""),             True),
        "apellido":    (p.get("apellido",""),            True),
        "diagnostico": (p.get("diagnostico",""),         False),
        "alergias":    (p.get("alergias",""),            False),
        "medicamentos":(p.get("medicamentos",""),        False),
        "telefono":    (p.get("telefono_emergencia",""), False),
    })
    if p.get("fecha_nacimiento"):
        ok, m = val("fecha", p["fecha_nacimiento"])
        if not ok: errores.append(m)
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("pacientes"))

    nuevo_id = qdb("SELECT nextval('seq_pacientes') AS id", one=True)["ID"]
    edb("INSERT INTO PACIENTES(ID,NOMBRE,APELLIDO,FECHA_NACIMIENTO,DIAGNOSTICO,"
        "ALERGIAS,MEDICAMENTOS,CONTACTO_EMERGENCIA,TELEFONO_EMERGENCIA) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (nuevo_id, p["nombre"], p["apellido"], p.get("fecha_nacimiento") or None,
         p.get("diagnostico",""), p.get("alergias",""), p.get("medicamentos",""),
         p.get("contacto_emergencia",""), p.get("telefono_emergencia","")))
    registrar_auditoria("CREAR", "PACIENTES", nuevo_id, f"Paciente registrado: {p['nombre']} {p['apellido']}")
    flash("Paciente registrado correctamente", "success")
    return redirect(url_for("pacientes"))

@app.route("/pacientes/<int:pid>")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def expediente(pid):
    pac = qdb("SELECT * FROM PACIENTES WHERE ID=%s", (pid,), one=True)
    if not pac: abort(404)
    historial = qdb(
        "SELECT S.*, U.NOMBRE||' '||U.APELLIDO TERAPEUTA, E.NOMBRE EQUINO "
        "FROM SESIONES S JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "WHERE S.PACIENTE_ID=%s ORDER BY S.FECHA_HORA DESC", (pid,)
    )
    fams = qdb("SELECT * FROM FAMILIARES WHERE PACIENTE_ID=%s", (pid,))

    total_sesiones = len(historial or [])
    realizadas = sum(1 for h in (historial or []) if h["ESTADO"] == "REALIZADA")

    return render_template("expediente.html", paciente=pac,
                           historial=historial, familiares=fams,
                           total_sesiones=total_sesiones, realizadas=realizadas)

@app.route("/pacientes/<int:pid>/pdf")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def expediente_pdf(pid):
    pac = qdb("SELECT * FROM PACIENTES WHERE ID=%s", (pid,), one=True)
    if not pac: abort(404)
    historial = qdb(
        "SELECT S.*, U.NOMBRE||' '||U.APELLIDO TERAPEUTA, E.NOMBRE EQUINO "
        "FROM SESIONES S JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "WHERE S.PACIENTE_ID=%s ORDER BY S.FECHA_HORA DESC", (pid,)
    )
    fams = qdb("SELECT * FROM FAMILIARES WHERE PACIENTE_ID=%s", (pid,))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            rightMargin=0.7*inch, leftMargin=0.7*inch,
                            topMargin=0.7*inch, bottomMargin=0.7*inch)
    ss = getSampleStyleSheet()
    AZUL_OSC = colors.HexColor("#1e3a5f")
    AZUL_MED = colors.HexColor("#2c5f8a")
    GRIS_CLR = colors.HexColor("#f4f6f8")
    ROJO = colors.HexColor("#a73838")

    t_t = ParagraphStyle("T", parent=ss["Title"], textColor=AZUL_OSC, fontSize=20, spaceAfter=2, fontName="Helvetica-Bold")
    t_s = ParagraphStyle("S", parent=ss["Normal"], textColor=colors.HexColor("#5a6678"), fontSize=10, spaceAfter=10)
    t_h = ParagraphStyle("H", parent=ss["Heading2"], textColor=AZUL_OSC, fontSize=13, spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")
    t_n = ParagraphStyle("N", parent=ss["Normal"], textColor=colors.HexColor("#3d4a5c"), fontSize=10)
    t_f = ParagraphStyle("F", parent=ss["Normal"], textColor=colors.grey, fontSize=8)

    story = [
        Paragraph("CUADRA ERRE", t_t),
        Paragraph("Expediente Clinico Confidencial", t_s),
        HRFlowable(width="100%", thickness=1.5, color=AZUL_OSC), Spacer(1,14),
    ]

    story.append(Paragraph("Datos del Paciente", t_h))
    datos = [
        ["Nombre completo", f"{pac['NOMBRE']} {pac['APELLIDO']}"],
        ["Fecha de nacimiento", str(pac["FECHA_NACIMIENTO"] or "No registrada")],
        ["Diagnostico", pac["DIAGNOSTICO"] or "No registrado"],
        ["Alergias", pac["ALERGIAS"] or "Ninguna registrada"],
        ["Medicamentos", pac["MEDICAMENTOS"] or "Ninguno registrado"],
        ["Contacto de emergencia", f"{pac['CONTACTO_EMERGENCIA'] or '-'} ({pac['TELEFONO_EMERGENCIA'] or '-'})"],
    ]
    t_datos = Table(datos, colWidths=[2*inch, 4.5*inch])
    t_datos.setStyle(TableStyle([
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9.5),
        ("TEXTCOLOR",(0,0),(0,-1),AZUL_OSC),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white,GRIS_CLR]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d4d9e0")),
        ("PADDING",(0,0),(-1,-1),7),("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    story.append(t_datos)
    story.append(Spacer(1,16))

    story.append(Paragraph("Familiares Registrados", t_h))
    if fams:
        fam_rows = [["Nombre", "Parentesco", "Telefono", "Email"]] + [
            [f"{f['NOMBRE']} {f['APELLIDO']}", f["PARENTESCO"] or "-", f["TELEFONO"] or "-", f["EMAIL"] or "-"]
            for f in fams
        ]
        t_fam = Table(fam_rows, colWidths=[2*inch, 1.3*inch, 1.3*inch, 1.9*inch])
        t_fam.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),AZUL_MED),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRIS_CLR]),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d4d9e0")),("PADDING",(0,0),(-1,-1),6),
        ]))
        story.append(t_fam)
    else:
        story.append(Paragraph("Sin familiares registrados.", t_n))
    story.append(Spacer(1,16))

    total = len(historial or [])
    realizadas = sum(1 for h in (historial or []) if h["ESTADO"] == "REALIZADA")
    story.append(Paragraph("Resumen Clinico", t_h))
    res = Table([["Total de sesiones", str(total)], ["Sesiones realizadas", str(realizadas)]],
                colWidths=[3*inch, 2*inch])
    res.setStyle(TableStyle([
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9.5),
        ("TEXTCOLOR",(0,0),(0,-1),AZUL_OSC),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white,GRIS_CLR]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d4d9e0")),("PADDING",(0,0),(-1,-1),6),
    ]))
    story.append(res)
    story.append(Spacer(1,16))

    story.append(Paragraph("Historial de Sesiones", t_h))
    if historial:
        hist_rows = [["Fecha", "Terapeuta", "Equino", "Estado", "Notas"]]
        for h in historial:
            notas = (h["NOTAS_SESION"] or "")[:60]
            hist_rows.append([
                str(h["FECHA_HORA"])[:16], h["TERAPEUTA"], h["EQUINO"] or "-",
                h["ESTADO"], notas
            ])
        t_hist = Table(hist_rows, colWidths=[1.2*inch, 1.3*inch, 1.0*inch, 1.0*inch, 2*inch])
        t_hist.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),AZUL_OSC),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRIS_CLR]),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d4d9e0")),("PADDING",(0,0),(-1,-1),5),
        ]))
        story.append(t_hist)
    else:
        story.append(Paragraph("Sin sesiones registradas.", t_n))

    story.append(Spacer(1,24))
    story.append(Paragraph(
        f"Documento confidencial generado el {date.today().strftime('%d/%m/%Y')} - "
        f"Sistema de Gestion Clinica Cuadra Erre - Los Cabos, BCS", t_f))

    doc.build(story)
    buf.seek(0)
    registrar_auditoria("EXPORTAR", "PACIENTES", pid, f"Expediente PDF generado: {pac['NOMBRE']} {pac['APELLIDO']}")
    nombre_archivo = f"expediente_{pac['NOMBRE']}_{pac['APELLIDO']}.pdf".replace(" ", "_")
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=nombre_archivo)

@app.route("/pacientes/<int:pid>/editar", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def editar_paciente(pid):
    p = request.form
    errores = val_fields({
        "nombre":      (p.get("nombre",""),      True),
        "apellido":    (p.get("apellido",""),     True),
        "diagnostico": (p.get("diagnostico",""),  False),
        "alergias":    (p.get("alergias",""),     False),
        "medicamentos":(p.get("medicamentos",""), False),
    })
    if p.get("fecha_nacimiento"):
        ok, m = val("fecha", p["fecha_nacimiento"])
        if not ok: errores.append(m)
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("expediente", pid=pid))
    edb("UPDATE PACIENTES SET NOMBRE=%s,APELLIDO=%s,FECHA_NACIMIENTO=%s,"
        "DIAGNOSTICO=%s,ALERGIAS=%s,MEDICAMENTOS=%s,"
        "CONTACTO_EMERGENCIA=%s,TELEFONO_EMERGENCIA=%s WHERE ID=%s",
        (p["nombre"], p["apellido"], p.get("fecha_nacimiento") or None,
         p.get("diagnostico",""), p.get("alergias",""), p.get("medicamentos",""),
         p.get("contacto_emergencia",""), p.get("telefono_emergencia",""), pid))
    registrar_auditoria("ACTUALIZAR", "PACIENTES", pid, "Datos del paciente actualizados")
    flash("Paciente actualizado correctamente", "success")
    return redirect(url_for("expediente", pid=pid))

@app.route("/pacientes/<int:pid>/familiar/nuevo", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR")
def nuevo_familiar(pid):
    p = request.form
    errores = val_fields({
        "nombre":     (p.get("nombre",""),    True),
        "apellido":   (p.get("apellido",""),  True),
        "parentesco": (p.get("parentesco",""),False),
        "telefono":   (p.get("telefono",""),  False),
        "email":      (p.get("email",""),     False),
    })
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("expediente", pid=pid))
    nuevo_id = qdb("SELECT nextval('seq_familiares') AS id", one=True)["ID"]
    edb("INSERT INTO FAMILIARES(ID,PACIENTE_ID,NOMBRE,APELLIDO,PARENTESCO,TELEFONO,EMAIL) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s)",
        (nuevo_id, pid, p["nombre"], p["apellido"], p.get("parentesco",""),
         p.get("telefono",""), p.get("email","")))
    registrar_auditoria("CREAR", "FAMILIARES", nuevo_id, f"Familiar agregado: {p['nombre']} {p['apellido']}")
    flash("Familiar registrado correctamente", "success")
    return redirect(url_for("expediente", pid=pid))

# ============================================================
#  EQUINOS
# ============================================================
@app.route("/equinos")
@login_req
def equinos():
    lista = qdb("SELECT * FROM EQUINOS WHERE ACTIVO=1 ORDER BY NOMBRE")
    regs  = qdb(
        "SELECT R.*, E.NOMBRE EQ_NOM, U.NOMBRE ENC "
        "FROM REGISTRO_EQUINO R JOIN EQUINOS E ON R.EQUINO_ID=E.ID "
        "LEFT JOIN USUARIOS U ON R.ENCARGADO_ID=U.ID "
        "WHERE CAST(R.FECHA AS DATE)=%s", (date.today().isoformat(),)
    )
    mant_pendiente = (qdb(
        "SELECT COUNT(*) C FROM MANTENIMIENTO_EQUINO WHERE ESTADO='PENDIENTE' AND FECHA_PROGRAMADA <= %s",
        ((date.today()+timedelta(days=7)).isoformat(),), one=True) or {}).get("C", 0)
    return render_template("equinos.html", equinos=lista, registros=regs, mant_pendiente=mant_pendiente)

@app.route("/equinos/<int:eid>/detalle")
@login_req
def equino_detalle(eid):
    eq = qdb("SELECT * FROM EQUINOS WHERE ID=%s", (eid,), one=True)
    if not eq: abort(404)

    historial = qdb(
        "SELECT R.*, U.NOMBRE||' '||U.APELLIDO ENCARGADO "
        "FROM REGISTRO_EQUINO R LEFT JOIN USUARIOS U ON R.ENCARGADO_ID=U.ID "
        "WHERE R.EQUINO_ID=%s ORDER BY R.FECHA DESC LIMIT 60", (eid,)
    )
    sesiones_recientes = qdb(
        "SELECT S.FECHA_HORA, S.ESTADO, P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "WHERE S.EQUINO_ID=%s ORDER BY S.FECHA_HORA DESC LIMIT 20", (eid,)
    )
    mantenimientos = qdb(
        "SELECT M.*, U.NOMBRE||' '||U.APELLIDO REGISTRADO "
        "FROM MANTENIMIENTO_EQUINO M LEFT JOIN USUARIOS U ON M.REGISTRADO_POR=U.ID "
        "WHERE M.EQUINO_ID=%s ORDER BY M.FECHA_PROGRAMADA DESC LIMIT 30", (eid,)
    )
    total_sesiones = (qdb("SELECT COUNT(*) C FROM SESIONES WHERE EQUINO_ID=%s", (eid,), one=True) or {}).get("C", 0)

    return render_template("equino_detalle.html", eq=eq, historial=historial,
                           sesiones=sesiones_recientes, mantenimientos=mantenimientos,
                           total_sesiones=total_sesiones)

@app.route("/equinos/agenda")
@login_req
def equinos_agenda():
    DIAS_ES = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    MESES_ES = ["enero","febrero","marzo","abril","mayo","junio","julio",
                "agosto","septiembre","octubre","noviembre","diciembre"]

    hoy = date.today()
    dias = []
    for i in range(7):
        d = hoy + timedelta(days=i)
        sesiones = qdb(
            "SELECT S.FECHA_HORA, S.ESTADO, P.NOMBRE||' '||P.APELLIDO PACIENTE, "
            "E.ID EQUINO_ID, E.NOMBRE EQUINO, E.ESTADO ESTADO_EQUINO "
            "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
            "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
            "WHERE CAST(S.FECHA_HORA AS DATE)=%s AND S.ESTADO != 'CANCELADA' "
            "ORDER BY S.FECHA_HORA", (d.isoformat(),)
        )
        for s in (sesiones or []):
            s["HORA"] = str(s["FECHA_HORA"])[11:16]
        nombre_dia = f"{DIAS_ES[d.weekday()]}, {d.day} de {MESES_ES[d.month-1]}"
        dias.append({"fecha": d, "nombre_dia": nombre_dia, "sesiones": sesiones or []})
    return render_template("equinos_agenda.html", dias=dias)

@app.route("/mantenimiento")
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS", "COORDINADOR")
def mantenimiento():
    filtro = request.args.get("estado", "")
    base_sql = (
        "SELECT M.*, E.NOMBRE EQUINO "
        "FROM MANTENIMIENTO_EQUINO M JOIN EQUINOS E ON M.EQUINO_ID=E.ID WHERE 1=1 "
    )
    args = []
    if filtro:
        base_sql += "AND M.ESTADO=%s "
        args.append(filtro)
    base_sql += "ORDER BY M.FECHA_PROGRAMADA ASC"
    lista = qdb(base_sql, tuple(args))

    vencidos = sum(1 for m in (lista or []) if m["ESTADO"]=="PENDIENTE" and m["FECHA_PROGRAMADA"] < date.today())
    proximos = sum(1 for m in (lista or []) if m["ESTADO"]=="PENDIENTE" and date.today() <= m["FECHA_PROGRAMADA"] <= date.today()+timedelta(days=7))

    equinos_list = qdb("SELECT ID, NOMBRE FROM EQUINOS WHERE ACTIVO=1 ORDER BY NOMBRE")
    return render_template("mantenimiento.html", mantenimientos=lista, equinos=equinos_list,
                           filtro=filtro, vencidos=vencidos, proximos=proximos,
                           hoy_iso=date.today().isoformat())

@app.route("/mantenimiento/nuevo", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS")
def nuevo_mantenimiento():
    p = request.form
    if not p.get("equino_id") or not p.get("tipo") or not p.get("fecha_programada"):
        flash("Equino, tipo y fecha son obligatorios", "danger")
        return redirect(url_for("mantenimiento"))

    ok, m = val("fecha", p["fecha_programada"])
    if not ok:
        flash(m, "danger")
        return redirect(url_for("mantenimiento"))

    nuevo_id = qdb("SELECT nextval('seq_mantenimiento') AS id", one=True)["ID"]
    edb(
        "INSERT INTO MANTENIMIENTO_EQUINO(ID,EQUINO_ID,TIPO,FECHA_PROGRAMADA,ESTADO,NOTAS,REGISTRADO_POR) "
        "VALUES(%s,%s,%s,%s,'PENDIENTE',%s,%s)",
        (nuevo_id, p["equino_id"], p["tipo"], p["fecha_programada"], p.get("notas",""), session["uid"])
    )
    registrar_auditoria("CREAR", "MANTENIMIENTO_EQUINO", nuevo_id, f"Mantenimiento programado: {p['tipo']}")
    flash("Mantenimiento programado correctamente", "success")
    return redirect(url_for("mantenimiento"))

@app.route("/mantenimiento/<int:mid>/completar", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS")
def completar_mantenimiento(mid):
    notas = request.form.get("notas_completar", "")
    edb(
        "UPDATE MANTENIMIENTO_EQUINO SET ESTADO='COMPLETADO', FECHA_REALIZADA=CURRENT_DATE, NOTAS=%s WHERE ID=%s",
        (notas, mid)
    )
    registrar_auditoria("ACTUALIZAR", "MANTENIMIENTO_EQUINO", mid, "Mantenimiento marcado como completado")
    flash("Mantenimiento marcado como completado", "success")
    return redirect(url_for("mantenimiento"))

@app.route("/equinos/nuevo", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS")
def nuevo_equino():
    p = request.form
    errores = val_fields({
        "eq_nombre": (p.get("nombre",""),     True),
        "raza":      (p.get("raza",""),       False),
        "color":     (p.get("color",""),      False),
        "notas":     (p.get("notas_salud",""),False),
    })
    if p.get("edad"):
        ok, m = val("edad", p["edad"])
        if not ok: errores.append(m)
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("equinos"))
    nuevo_id = qdb("SELECT nextval('seq_equinos') AS id", one=True)["ID"]
    edb("INSERT INTO EQUINOS(ID,NOMBRE,RAZA,EDAD,COLOR,ESTADO,NOTAS_SALUD) "
        "VALUES(%s,%s,%s,%s,%s,'DISPONIBLE',%s)",
        (nuevo_id, p["nombre"], p.get("raza",""), p.get("edad") or None,
         p.get("color",""), p.get("notas_salud","")))
    registrar_auditoria("CREAR", "EQUINOS", nuevo_id, f"Equino registrado: {p['nombre']}")
    flash("Equino registrado correctamente", "success")
    return redirect(url_for("equinos"))

@app.route("/equinos/registro", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS", "COORDINADOR")
def registro_equino():
    p   = request.form
    eid = p.get("equino_id")
    if not eid:
        flash("Selecciona un equino", "danger")
        return redirect(url_for("equinos"))
    notas = p.get("observaciones","")
    ok, m = val("notas", notas, False)
    if not ok:
        flash(m, "danger"); return redirect(url_for("equinos"))
    estado_map = {"1":"DISPONIBLE","2":"PRECAUCION","3":"DESCANSO","4":"VETERINARIO"}
    sem = p.get("semaforo","1")

    nuevo_id = qdb("SELECT nextval('seq_registro_equino') AS id", one=True)["ID"]
    edb("INSERT INTO REGISTRO_EQUINO(ID,EQUINO_ID,FECHA,ENCARGADO_ID,"
        "ESTADO_FISICO,ESTADO_ANIMO,DISPONIBLE,OBSERVACIONES) "
        "VALUES(%s,%s,CURRENT_DATE,%s,%s,%s,%s,%s)",
        (nuevo_id, eid, session["uid"], p.get("estado_fisico","BUENO"),
         p.get("estado_animo","TRANQUILO"), 1 if sem=="1" else 0, notas))
    edb("UPDATE EQUINOS SET ESTADO=%s WHERE ID=%s", (estado_map.get(sem,"DISPONIBLE"), eid))
    registrar_auditoria("ACTUALIZAR", "EQUINOS", eid, f"Estado actualizado a {estado_map.get(sem)}")

    if sem in ("3", "4"):
        afectadas = qdb(
            "SELECT ID FROM SESIONES WHERE EQUINO_ID=%s AND ESTADO='PROGRAMADA' AND FECHA_HORA >= CURRENT_TIMESTAMP",
            (eid,)
        )
        if afectadas:
            flash(f"Atencion: {len(afectadas)} sesion(es) futuras tienen asignado este equino y requieren reasignacion", "warning")

    flash("Estado del equino actualizado correctamente", "success")
    return redirect(url_for("equinos"))

@app.route("/equinos/<int:eid>/editar", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR", "ENCARGADO_EQUINOS")
def editar_equino(eid):
    p = request.form
    errores = val_fields({
        "eq_nombre": (p.get("nombre",""),     True),
        "raza":      (p.get("raza",""),       False),
        "color":     (p.get("color",""),      False),
        "notas":     (p.get("notas_salud",""),False),
    })
    if p.get("edad"):
        ok, m = val("edad", p["edad"])
        if not ok: errores.append(m)
    if p.get("ultima_revision"):
        ok, m = val("fecha", p["ultima_revision"])
        if not ok: errores.append(m)
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("equinos"))
    edb("UPDATE EQUINOS SET NOMBRE=%s,RAZA=%s,EDAD=%s,COLOR=%s,NOTAS_SALUD=%s,ULTIMA_REVISION=%s WHERE ID=%s",
        (p["nombre"], p.get("raza",""), p.get("edad") or None,
         p.get("color",""), p.get("notas_salud",""),
         p.get("ultima_revision") or None, eid))
    registrar_auditoria("ACTUALIZAR", "EQUINOS", eid, "Datos del equino actualizados")
    flash("Equino actualizado correctamente", "success")
    return redirect(url_for("equinos"))

# ============================================================
#  USUARIOS
# ============================================================
@app.route("/usuarios")
@login_req
@rol_req("ADMINISTRADOR")
def usuarios():
    lista = qdb("SELECT U.*, R.NOMBRE ROL_NOM FROM USUARIOS U JOIN ROLES R ON U.ROL_ID=R.ID ORDER BY U.NOMBRE")
    roles = qdb("SELECT * FROM ROLES ORDER BY ID")
    return render_template("usuarios.html", usuarios=lista, roles=roles)

@app.route("/usuarios/nuevo", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR")
def nuevo_usuario():
    p     = request.form
    uname = p.get("username","").strip()
    pwd   = p.get("password","")
    errores = val_fields({
        "user_create": (uname,               True),
        "password":    (pwd,                 True),
        "nombre":      (p.get("nombre",""),  True),
        "apellido":    (p.get("apellido",""),False),
        "email":       (p.get("email",""),   False),
        "telefono":    (p.get("telefono",""),False),
    })
    if not p.get("rol_id"): errores.append("Rol: obligatorio")
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("usuarios"))
    if qdb("SELECT ID FROM USUARIOS WHERE LOWER(USERNAME)=%s", (uname.lower(),), one=True):
        flash("El nombre de usuario ya existe", "danger")
        return redirect(url_for("usuarios"))

    nuevo_id = qdb("SELECT nextval('seq_usuarios') AS id", one=True)["ID"]
    edb("INSERT INTO USUARIOS(ID,USERNAME,PASSWORD_HASH,NOMBRE,APELLIDO,EMAIL,TELEFONO,ROL_ID,ACTIVO) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,1)",
        (nuevo_id, uname, generate_password_hash(pwd),
         p["nombre"], p.get("apellido",""),
         p.get("email",""), p.get("telefono",""), p["rol_id"]))
    registrar_auditoria("CREAR", "USUARIOS", nuevo_id, f"Usuario creado: {uname}")
    flash(f"Usuario '{uname}' creado correctamente", "success")
    return redirect(url_for("usuarios"))

@app.route("/usuarios/<int:uid>/editar", methods=["POST"])
@login_req
@rol_req("ADMINISTRADOR")
def editar_usuario(uid):
    p = request.form
    errores = val_fields({
        "nombre":   (p.get("nombre",""),   True),
        "apellido": (p.get("apellido",""), False),
        "email":    (p.get("email",""),    False),
        "telefono": (p.get("telefono",""), False),
    })
    if p.get("password"):
        ok, m = val("password", p["password"])
        if not ok: errores.append(m)
    if errores:
        flash(" | ".join(errores), "danger")
        return redirect(url_for("usuarios"))
    if p.get("password"):
        edb("UPDATE USUARIOS SET NOMBRE=%s,APELLIDO=%s,EMAIL=%s,TELEFONO=%s,ROL_ID=%s,ACTIVO=%s,PASSWORD_HASH=%s WHERE ID=%s",
            (p["nombre"],p.get("apellido",""),p.get("email",""),p.get("telefono",""),
             p["rol_id"],int(p.get("activo",1)),generate_password_hash(p["password"]),uid))
    else:
        edb("UPDATE USUARIOS SET NOMBRE=%s,APELLIDO=%s,EMAIL=%s,TELEFONO=%s,ROL_ID=%s,ACTIVO=%s WHERE ID=%s",
            (p["nombre"],p.get("apellido",""),p.get("email",""),p.get("telefono",""),
             p["rol_id"],int(p.get("activo",1)),uid))
    registrar_auditoria("ACTUALIZAR", "USUARIOS", uid, "Datos de usuario actualizados")
    flash("Usuario actualizado correctamente", "success")
    return redirect(url_for("usuarios"))

@app.route("/perfil", methods=["GET", "POST"])
@login_req
def perfil():
    usuario = qdb("SELECT * FROM USUARIOS WHERE ID=%s", (session["uid"],), one=True)
    if request.method == "POST":
        p = request.form
        pwd_actual = p.get("password_actual", "")
        pwd_nueva  = p.get("password_nueva", "")

        if not check_password_hash(usuario["PASSWORD_HASH"], pwd_actual):
            flash("La contrasena actual es incorrecta", "danger")
            return redirect(url_for("perfil"))

        ok, m = val("password", pwd_nueva)
        if not ok:
            flash(m, "danger")
            return redirect(url_for("perfil"))

        edb("UPDATE USUARIOS SET PASSWORD_HASH=%s WHERE ID=%s",
            (generate_password_hash(pwd_nueva), session["uid"]))
        registrar_auditoria("ACTUALIZAR", "USUARIOS", session["uid"], "Cambio de contrasena propia")
        flash("Contrasena actualizada correctamente", "success")
        return redirect(url_for("perfil"))

    return render_template("perfil.html", usuario=usuario)

@app.route("/configuracion", methods=["GET", "POST"])
@login_req
@rol_req("ADMINISTRADOR")
def configuracion():
    cfg = qdb("SELECT * FROM CONFIGURACION LIMIT 1", one=True)
    if request.method == "POST":
        p = request.form
        errores = val_fields({
            "nombre": (p.get("nombre_centro",""), True),
            "telefono": (p.get("telefono",""), False),
            "email": (p.get("email_contacto",""), False),
        })
        if errores:
            flash(" | ".join(errores), "danger")
            return redirect(url_for("configuracion"))

        if cfg:
            edb(
                "UPDATE CONFIGURACION SET NOMBRE_CENTRO=%s,TELEFONO=%s,DIRECCION=%s,"
                "EMAIL_CONTACTO=%s,RECORDATORIO_24H=%s,HORAS_ANTICIPACION=%s WHERE ID=%s",
                (p.get("nombre_centro","Cuadra Erre"), p.get("telefono",""), p.get("direccion",""),
                 p.get("email_contacto",""),
                 1 if p.get("recordatorio_24h") == "1" else 0,
                 p.get("horas_anticipacion", "24"), cfg["ID"])
            )
        else:
            nuevo_id = qdb("SELECT nextval('seq_configuracion') AS id", one=True)["ID"]
            edb(
                "INSERT INTO CONFIGURACION(ID,NOMBRE_CENTRO,TELEFONO,DIRECCION,EMAIL_CONTACTO,"
                "RECORDATORIO_24H,HORAS_ANTICIPACION) VALUES(%s,%s,%s,%s,%s,%s,%s)",
                (nuevo_id, p.get("nombre_centro","Cuadra Erre"), p.get("telefono",""),
                 p.get("direccion",""), p.get("email_contacto",""),
                 1 if p.get("recordatorio_24h") == "1" else 0, p.get("horas_anticipacion", "24"))
            )
        registrar_auditoria("ACTUALIZAR", "CONFIGURACION", cfg["ID"] if cfg else None, "Configuracion del sistema actualizada")
        flash("Configuracion guardada correctamente", "success")
        return redirect(url_for("configuracion"))

    return render_template("configuracion.html", cfg=cfg or {})

# ============================================================
#  NOTIFICACIONES
# ============================================================
@app.route("/notificaciones")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def notificaciones():
    lista = qdb(
        "SELECT N.*, F.NOMBRE||' '||F.APELLIDO FAMILIAR, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE "
        "FROM NOTIFICACIONES N JOIN FAMILIARES F ON N.FAMILIAR_ID=F.ID "
        "JOIN PACIENTES P ON F.PACIENTE_ID=P.ID "
        "ORDER BY N.FECHA_CREACION DESC LIMIT 80"
    )
    return render_template("notificaciones.html", notificaciones=lista)

@app.route("/notificaciones/<int:nid>/marcar", methods=["POST"])
@login_req
def marcar_notif(nid):
    edb("UPDATE NOTIFICACIONES SET ENVIADA=1, FECHA_ENVIO=CURRENT_TIMESTAMP WHERE ID=%s", (nid,))
    return jsonify({"ok": True})

# ============================================================
#  AUDITORIA
# ============================================================
@app.route("/auditoria")
@login_req
@rol_req("ADMINISTRADOR")
def auditoria():
    registros = qdb(
        "SELECT A.*, U.USERNAME, U.NOMBRE||' '||U.APELLIDO NOMBRE_USUARIO "
        "FROM AUDITORIA A LEFT JOIN USUARIOS U ON A.USUARIO_ID=U.ID "
        "ORDER BY A.FECHA DESC LIMIT 200"
    )
    return render_template("auditoria.html", registros=registros)

# ============================================================
#  REPORTE SEMANAL + PDF
# ============================================================
@app.route("/reporte")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR")
def reporte():
    hoy = date.today()
    ini = hoy - timedelta(days=hoy.weekday())
    fin = ini + timedelta(days=6)
    return render_template("reporte.html", inicio=ini, fin=fin)

@app.route("/reporte/pdf")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR")
def reporte_pdf():
    desde = request.args.get("desde", (date.today()-timedelta(days=7)).isoformat())
    hasta = request.args.get("hasta", date.today().isoformat())

    sesiones = qdb(
        "SELECT S.FECHA_HORA, S.ESTADO, S.DURACION_MIN, S.NOTAS_SESION, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, E.NOMBRE EQUINO "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "WHERE CAST(S.FECHA_HORA AS DATE) BETWEEN %s AND %s "
        "ORDER BY S.FECHA_HORA", (desde, hasta)
    )
    carga = qdb(
        "SELECT U.NOMBRE||' '||U.APELLIDO TERAPEUTA, COUNT(*) TOTAL "
        "FROM SESIONES S JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "WHERE CAST(S.FECHA_HORA AS DATE) BETWEEN %s AND %s AND S.ESTADO='REALIZADA' "
        "GROUP BY U.NOMBRE, U.APELLIDO ORDER BY TOTAL DESC", (desde, hasta)
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            rightMargin=0.7*inch, leftMargin=0.7*inch,
                            topMargin=0.7*inch, bottomMargin=0.7*inch)
    ss = getSampleStyleSheet()
    AZUL_OSC = colors.HexColor("#1e3a5f")
    AZUL_MED = colors.HexColor("#2c5f8a")
    GRIS_CLR = colors.HexColor("#f4f6f8")

    t_t = ParagraphStyle("T", parent=ss["Title"], textColor=AZUL_OSC, fontSize=20, spaceAfter=2, fontName="Helvetica-Bold")
    t_s = ParagraphStyle("S", parent=ss["Normal"], textColor=colors.HexColor("#5a6678"), fontSize=10, spaceAfter=10)
    t_h = ParagraphStyle("H", parent=ss["Heading2"], textColor=AZUL_OSC, fontSize=13, spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold")
    t_f = ParagraphStyle("F", parent=ss["Normal"], textColor=colors.grey, fontSize=8)

    real = sum(1 for s in sesiones if s["ESTADO"]=="REALIZADA")
    prog = sum(1 for s in sesiones if s["ESTADO"]=="PROGRAMADA")
    canc = sum(1 for s in sesiones if s["ESTADO"]=="CANCELADA")

    story = [
        Paragraph("CUADRA ERRE", t_t),
        Paragraph(f"Reporte Semanal de Operaciones &nbsp;|&nbsp; Periodo: {desde} al {hasta}", t_s),
        HRFlowable(width="100%", thickness=1.5, color=AZUL_OSC), Spacer(1,12),
    ]
    res = Table([["Metrica","Valor"],["Total de sesiones",str(len(sesiones))],
                 ["Sesiones realizadas",str(real)],["Sesiones programadas",str(prog)],
                 ["Sesiones canceladas",str(canc)]],
                colWidths=[3*inch,1.5*inch])
    res.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),AZUL_OSC),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[GRIS_CLR,colors.white]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#d4d9e0")),("PADDING",(0,0),(-1,-1),7),
        ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),("TEXTCOLOR",(0,1),(0,-1),AZUL_OSC),
    ]))
    story.append(res); story.append(Spacer(1,16))

    story.append(Paragraph("Detalle de Sesiones", t_h))
    if sesiones:
        hdr = ["Fecha/Hora","Paciente","Terapeuta","Equino","Estado","Min"]
        rows = [[str(s["FECHA_HORA"])[:16], s["PACIENTE"], s["TERAPEUTA"],
                 s["EQUINO"] or "-", s["ESTADO"], str(s["DURACION_MIN"] or 45)]
                for s in sesiones]
        t = Table([hdr]+rows, colWidths=[1.35*inch,1.4*inch,1.4*inch,1.0*inch,1.0*inch,0.5*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),AZUL_OSC),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRIS_CLR]),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d4d9e0")),
            ("ALIGN",(5,0),(5,-1),"CENTER"),("PADDING",(0,0),(-1,-1),5),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Sin sesiones registradas en este periodo.", ss["Normal"]))

    if carga:
        story.append(Spacer(1,16))
        story.append(Paragraph("Carga de Trabajo por Terapeuta", t_h))
        ct = Table([["Terapeuta","Sesiones realizadas"]]+
                   [[c["TERAPEUTA"],str(c["TOTAL"])] for c in carga],
                   colWidths=[3*inch,2*inch])
        ct.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),AZUL_MED),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#eef3f8")]),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#c5d2e3")),
            ("ALIGN",(1,0),(1,-1),"CENTER"),("PADDING",(0,0),(-1,-1),6),
        ]))
        story.append(ct)

    story.append(Spacer(1,24))
    story.append(Paragraph(f"Documento generado el {date.today().strftime('%d/%m/%Y')} - Sistema de Gestion Clinica Cuadra Erre - Los Cabos, BCS", t_f))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"reporte_cuadraerre_{desde}_{hasta}.pdf")

@app.route("/api/equinos")
@login_req
def api_equinos():
    return jsonify(qdb("SELECT ID, NOMBRE, ESTADO FROM EQUINOS WHERE ACTIVO=1 ORDER BY NOMBRE"))

# ============================================================
#  EXPORTAR A EXCEL
# ============================================================
@app.route("/pacientes/exportar")
@login_req
@rol_req("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA")
def exportar_pacientes():
    lista = qdb("SELECT * FROM PACIENTES WHERE ACTIVO=1 ORDER BY APELLIDO, NOMBRE")
    headers = ["Nombre", "Apellido", "Fecha Nacimiento", "Diagnostico", "Alergias",
               "Medicamentos", "Contacto Emergencia", "Telefono Emergencia"]
    rows = [[
        p["NOMBRE"], p["APELLIDO"], str(p["FECHA_NACIMIENTO"] or ""),
        p["DIAGNOSTICO"] or "", p["ALERGIAS"] or "", p["MEDICAMENTOS"] or "",
        p["CONTACTO_EMERGENCIA"] or "", p["TELEFONO_EMERGENCIA"] or ""
    ] for p in (lista or [])]
    registrar_auditoria("EXPORTAR", "PACIENTES", None, "Exportacion a Excel")
    return generar_excel("Pacientes", headers, rows, f"pacientes_cuadraerre_{date.today().isoformat()}.xlsx")

@app.route("/equinos/exportar")
@login_req
def exportar_equinos():
    lista = qdb("SELECT * FROM EQUINOS WHERE ACTIVO=1 ORDER BY NOMBRE")
    headers = ["Nombre", "Raza", "Edad", "Color", "Estado", "Notas de Salud", "Ultima Revision"]
    rows = [[
        e["NOMBRE"], e["RAZA"] or "", e["EDAD"] or "", e["COLOR"] or "",
        e["ESTADO"] or "", e["NOTAS_SALUD"] or "", str(e["ULTIMA_REVISION"] or "")
    ] for e in (lista or [])]
    registrar_auditoria("EXPORTAR", "EQUINOS", None, "Exportacion a Excel")
    return generar_excel("Equinos", headers, rows, f"equinos_cuadraerre_{date.today().isoformat()}.xlsx")

@app.route("/agenda/exportar")
@login_req
def exportar_sesiones():
    base_sql = (
        "SELECT S.FECHA_HORA, S.ESTADO, S.DURACION_MIN, "
        "P.NOMBRE||' '||P.APELLIDO PACIENTE, "
        "U.NOMBRE||' '||U.APELLIDO TERAPEUTA, "
        "E.NOMBRE EQUINO, A.NOMBRE AREA, S.OBJETIVOS, S.NOTAS_SESION "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "JOIN USUARIOS U ON S.TERAPEUTA_ID=U.ID "
        "LEFT JOIN EQUINOS E ON S.EQUINO_ID=E.ID "
        "LEFT JOIN AREAS A ON S.AREA_ID=A.ID WHERE 1=1 "
    )
    args = []
    if session["rol"] == "TERAPEUTA":
        base_sql += "AND S.TERAPEUTA_ID=%s "
        args.append(session["uid"])
    base_sql += "ORDER BY S.FECHA_HORA DESC"

    lista = qdb(base_sql, tuple(args))
    headers = ["Fecha y Hora", "Paciente", "Terapeuta", "Equino", "Area",
               "Duracion (min)", "Estado", "Objetivos", "Notas"]
    rows = [[
        str(s["FECHA_HORA"])[:16], s["PACIENTE"], s["TERAPEUTA"], s["EQUINO"] or "",
        s["AREA"] or "", s["DURACION_MIN"] or 45, s["ESTADO"],
        s["OBJETIVOS"] or "", s["NOTAS_SESION"] or ""
    ] for s in (lista or [])]
    registrar_auditoria("EXPORTAR", "SESIONES", None, "Exportacion a Excel")
    return generar_excel("Sesiones", headers, rows, f"sesiones_cuadraerre_{date.today().isoformat()}.xlsx")

# ============================================================
#  BUSQUEDA GLOBAL
# ============================================================
@app.route("/api/buscar")
@login_req
def api_buscar():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify({"resultados": []})

    like = f"%{q.upper()}%"
    resultados = []

    if session["rol"] in ("ADMINISTRADOR", "COORDINADOR", "TERAPEUTA"):
        pacientes = qdb(
            "SELECT ID, NOMBRE, APELLIDO FROM PACIENTES WHERE ACTIVO=1 AND "
            "UPPER(NOMBRE||' '||APELLIDO) LIKE %s LIMIT 5", (like,)
        )
        for p in (pacientes or []):
            resultados.append({
                "tipo": "Paciente", "titulo": f"{p['NOMBRE']} {p['APELLIDO']}",
                "url": f"/pacientes/{p['ID']}"
            })

    equinos = qdb(
        "SELECT ID, NOMBRE FROM EQUINOS WHERE ACTIVO=1 AND UPPER(NOMBRE) LIKE %s LIMIT 5", (like,)
    )
    for e in (equinos or []):
        resultados.append({"tipo": "Equino", "titulo": e["NOMBRE"], "url": "/equinos"})

    if session["rol"] == "ADMINISTRADOR":
        usuarios = qdb(
            "SELECT ID, NOMBRE, APELLIDO, USERNAME FROM USUARIOS WHERE "
            "UPPER(NOMBRE||' '||APELLIDO||' '||USERNAME) LIKE %s LIMIT 5", (like,)
        )
        for u in (usuarios or []):
            resultados.append({
                "tipo": "Usuario", "titulo": f"{u['NOMBRE']} {u['APELLIDO'] or ''} (@{u['USERNAME']})",
                "url": "/usuarios"
            })

    sesiones_sql = (
        "SELECT S.ID, P.NOMBRE||' '||P.APELLIDO PACIENTE, S.FECHA_HORA "
        "FROM SESIONES S JOIN PACIENTES P ON S.PACIENTE_ID=P.ID "
        "WHERE UPPER(P.NOMBRE||' '||P.APELLIDO) LIKE %s "
    )
    sargs = [like]
    if session["rol"] == "TERAPEUTA":
        sesiones_sql += "AND S.TERAPEUTA_ID=%s "
        sargs.append(session["uid"])
    sesiones_sql += "ORDER BY S.FECHA_HORA DESC LIMIT 5"
    sesiones = qdb(sesiones_sql, tuple(sargs))
    for s in (sesiones or []):
        resultados.append({
            "tipo": "Sesion", "titulo": f"{s['PACIENTE']} - {str(s['FECHA_HORA'])[:16]}",
            "url": f"/sesion/{s['ID']}"
        })

    return jsonify({"resultados": resultados})

if __name__ == "__main__":
    app.run(debug=False)
